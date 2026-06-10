"""
Excel Exporter
Writes all school records to an Excel spreadsheet matching the template format.
"""

import csv
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.scraper import SchoolRecord

log = logging.getLogger(__name__)

# Column headers matching the provided template exactly
COLUMNS = [
    "School",
    "city",
    "category",
    "website",
    "description",
    "verification_status",
    "address",
    "phone",
]

ENRICHED_COLUMNS = [
    "founded",
    "ages",
    "students",
    "ratio",
    "school_type",
    "annual_fee",
    "tagline",
]


def export_to_excel(schools: list[SchoolRecord], output_path: str = "output/schools_data.xlsx"):
    """Export school data to Excel with formatting."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schools"

    # Header style
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    all_columns = COLUMNS + ENRICHED_COLUMNS

    # Write headers
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # Data rows
    row_fill_even = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, school in enumerate(schools, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        row_data = [
            school.name,
            school.city,
            school.category,
            school.website,
            school.description,
            school.verification_status,
            school.address,
            school.phone,
            school.founded,
            school.ages,
            school.students,
            school.ratio,
            school.school_type,
            school.annual_fee,
            school.tagline,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 60

    # Column widths
    widths = [30, 15, 20, 25, 60, 18, 40, 18, 10, 10, 12, 10, 30, 30, 50]
    for col_idx, width in enumerate(widths, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"Excel saved: {output_path}")
    print(f"Excel exported: {output_path}")
    return output_path


def export_missing_report(missing: list, output_path: str = "output/missing_report.xlsx"):
    """Export missing data report to Excel."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Data Report"

    headers = ["School Name", "Missing Field(s)", "Reason", "Location"]
    header_fill = PatternFill(start_color="C53030", end_color="C53030", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(missing, start=2):
        ws.cell(row=row_idx, column=1, value=record.get("School Name", ""))
        ws.cell(row=row_idx, column=2, value=record.get("Missing Field(s)", ""))
        ws.cell(row=row_idx, column=3, value=record.get("Reason", ""))
        ws.cell(row=row_idx, column=4, value=record.get("Location", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    log.info(f"Missing report saved: {output_path}")
    print(f"Missing report exported: {output_path}")
    return output_path


def export_to_csv(schools: list[SchoolRecord], output_path: str = "output/schools_data.csv"):
    """Also export as CSV for convenience."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    all_columns = COLUMNS + ENRICHED_COLUMNS
    field_map = {
        "School": "name", "city": "city", "category": "category",
        "website": "website", "description": "description",
        "verification_status": "verification_status", "address": "address",
        "phone": "phone", "founded": "founded", "ages": "ages",
        "students": "students", "ratio": "ratio", "school_type": "school_type",
        "annual_fee": "annual_fee", "tagline": "tagline",
    }
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_columns)
        writer.writeheader()
        for school in schools:
            writer.writerow({col: getattr(school, field_map[col], "") for col in all_columns})
    log.info(f"CSV saved: {output_path}")
    print(f"CSV exported: {output_path}")
    return output_path
